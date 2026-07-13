#!/usr/bin/env python3
"""Шаблон: генерация XLSX тест-кейсов для импорта в Test IT (15 колонок).

Копируй этот файл, заменяй данные CASES на свои и запускай.
"""

import os
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ---------------------------------------------------------------------------
# Данные тест-кейсов — ЗАМЕНИ НА СВОИ
# ---------------------------------------------------------------------------
BASE_SRC = "Путь_к_требованию"

CASES = [
    {
        "name": "TC-001 – Название тест-кейса",
        "desc": f"Источник: {BASE_SRC}\nТребование: Описание требования",
        "priority": "Высокий",
        "duration": "0h 5m 0s",
        "section": "Секция > Подсекция",
        "review": "Не пройдено",
        "type": "Позитивный",
        "preconditions": [
            "Предусловие 1",
            "Предусловие 2",
        ],
        "steps": [
            {
                "action": "Шаг 1 — действие пользователя",
                "expected": "Ожидаемый результат 1",
            },
            {
                "action": "Шаг 2 — действие пользователя",
                "expected": "Ожидаемый результат 2",
            },
        ],
    },
    # Добавляй дополнительные тест-кейсы по аналогии
]

# ---------------------------------------------------------------------------
# Стилизация
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

COLUMNS = [
    "Название",
    "Описание",
    "Статус",
    "Приоритет",
    "Действие",
    "Ожидаемый результат",
    "Действие предусловия",
    "Ожидаемый результат предусловия",
    "Действие постусловия",
    "Ожидаемый результат постусловия",
    "Продолжительность",
    "Секция",
    "Тестовые данные",
    "Ревью",
    "Вид",
]


def write_header_row(ws, row: int, d: dict):
    """Header-строка тест-кейса (колонки 1-15, заполнены только обязательные)."""
    vals = [
        d["name"],              # 1 Название
        d["desc"],              # 2 Описание
        "Готов",                # 3 Статус
        d["priority"],          # 4 Приоритет
        "",                     # 5 Действие
        "",                     # 6 Ожидаемый результат
        "",                     # 7 Действие предусловия
        "",                     # 8 Ожидаемый результат предусловия
        "",                     # 9 Действие постусловия
        "",                     # 10 Ожидаемый результат постусловия
        d["duration"],          # 11 Продолжительность
        d["section"],           # 12 Секция
        "",                     # 13 Тестовые данные
        d["review"],            # 14 Ревью
        d["type"],              # 15 Вид
    ]
    for col_idx, val in enumerate(vals, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        cell.border = THIN_BORDER


def write_row(ws, row: int, col: int, value: str):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = BODY_FONT
    cell.alignment = WRAP
    cell.border = THIN_BORDER


def write_case(ws: Workbook, start_row: int, case: dict) -> int:
    """Write one test case. Returns next available row number."""
    row = start_row

    # Header
    write_header_row(ws, row, case)
    row += 1

    # Preconditions
    for pc in case["preconditions"]:
        write_row(ws, row, 1, "")
        write_row(ws, row, 2, "")
        write_row(ws, row, 3, "")
        write_row(ws, row, 4, "")
        write_row(ws, row, 5, "")
        write_row(ws, row, 6, "")
        write_row(ws, row, 7, pc)
        write_row(ws, row, 8, "")
        write_row(ws, row, 9, "")
        write_row(ws, row, 10, "")
        write_row(ws, row, 11, "")
        write_row(ws, row, 12, "")
        write_row(ws, row, 13, "")
        write_row(ws, row, 14, "")
        write_row(ws, row, 15, "")
        row += 1

    # Steps + Expected results
    for step in case["steps"]:
        write_row(ws, row, 1, "")
        write_row(ws, row, 2, "")
        write_row(ws, row, 3, "")
        write_row(ws, row, 4, "")
        write_row(ws, row, 5, step["action"])
        write_row(ws, row, 6, step["expected"])
        write_row(ws, row, 7, "")
        write_row(ws, row, 8, "")
        write_row(ws, row, 9, "")
        write_row(ws, row, 10, "")
        write_row(ws, row, 11, "")
        write_row(ws, row, 12, "")
        write_row(ws, row, 13, "")
        write_row(ws, row, 14, "")
        write_row(ws, row, 15, "")
        row += 1

    # Empty separator row
    row += 1
    return row


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Тест-кейсы"

    # Column headers
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(name="Calibri", bold=True, size=11)
        cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        cell.alignment = WRAP
        cell.border = THIN_BORDER

    # Write all test cases
    row = 2
    for case in CASES:
        row = write_case(ws, row, case)

    # Column widths
    col_widths = {1: 50, 2: 60, 3: 10, 4: 12, 5: 55, 6: 65, 7: 55, 8: 10,
                  9: 55, 10: 10, 11: 14, 12: 30, 13: 14, 14: 12, 15: 12}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # Freeze top row
    ws.freeze_panes = "A2"

    # Save
    output_path = "/Users/redcollar/dev/TC_Output.xlsx"
    wb.save(output_path)
    time.sleep(2)

    import os as _os
    exists = _os.path.exists(output_path)
    size = _os.path.getsize(output_path) if exists else 0
    print(f"[{output_path}] exists={exists} size={size} bytes cases={len(CASES)}")


if __name__ == "__main__":
    main()
